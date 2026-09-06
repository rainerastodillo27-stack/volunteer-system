"""
Script to swap 'Expected Outcome' and 'Actual Outcome' columns in all testing documentation:
- 4 Word documents (.docx) in docs/testings/
- 2 Black-Box Excel workbooks (.xlsx) in docs/testings/
So that 'Actual Outcome' comes FIRST, followed by 'Expected Outcome' across ALL docs.
"""

import os
import shutil
import docx
import openpyxl

DOCS_DIR = os.path.abspath("docs/testings")
BACKUP_DIR = os.path.join(DOCS_DIR, "backup")

def backup_files():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    for f in os.listdir(DOCS_DIR):
        src = os.path.join(DOCS_DIR, f)
        if os.path.isfile(src) and (f.endswith(".docx") or f.endswith(".xlsx")):
            dest = os.path.join(BACKUP_DIR, f)
            if not os.path.exists(dest):
                shutil.copy2(src, dest)
                print(f"[BACKUP] Backed up {f} -> backup/{f}")
            else:
                print(f"[BACKUP] Existing backup found for {f}")

def update_docx(filename):
    filepath = os.path.join(DOCS_DIR, filename)
    print(f"\n[DOCX] Processing {filename}...")
    doc = docx.Document(filepath)
    swapped_tables = 0

    for tbl_idx, table in enumerate(doc.tables):
        if len(table.rows) == 0 or len(table.rows[0].cells) < 6:
            continue

        header_4 = table.rows[0].cells[4].text.strip().lower()
        header_5 = table.rows[0].cells[5].text.strip().lower()

        # Check if column 4 is Expected Outcome and column 5 is Actual Outcome
        if "expected" in header_4 and "actual" in header_5:
            for row in table.rows:
                tc4 = row.cells[4]._tc
                tc5 = row.cells[5]._tc
                idx4 = row._tr.index(tc4)
                row._tr.remove(tc5)
                row._tr.insert(idx4, tc5)
            swapped_tables += 1

    doc.save(filepath)
    print(f"[DOCX] Successfully swapped {swapped_tables} tables in {filename}")
    return swapped_tables

def update_xlsx(filename):
    filepath = os.path.join(DOCS_DIR, filename)
    print(f"\n[XLSX] Processing {filename}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Check header row
    h4 = str(ws.cell(row=1, column=4).value or "").strip().lower()
    h5 = str(ws.cell(row=1, column=5).value or "").strip().lower()

    if "expected" in h4 and "actual" in h5:
        row_count = ws.max_row
        for r in range(1, row_count + 1):
            c4 = ws.cell(row=r, column=4)
            c5 = ws.cell(row=r, column=5)
            val4 = c4.value
            val5 = c5.value
            c4.value = val5
            c5.value = val4
        wb.save(filepath)
        print(f"[XLSX] Successfully swapped columns 4 and 5 across {row_count} rows in {filename}")
    else:
        print(f"[XLSX] Header already in order or not matching: Col 4='{h4}', Col 5='{h5}'")

def main():
    backup_files()

    docx_files = [
        "Alpha_Blackbox_Test_Report.docx",
        "Beta_Blackbox_Test_Report.docx",
        "Alpha_Whitebox_Test_Report.docx",
        "Beta_Whitebox_Test_Report.docx"
    ]

    xlsx_files = [
        "alpha_blackbox_test_results.xlsx",
        "beta_blackbox_test_results.xlsx"
    ]

    for f in docx_files:
        update_docx(f)

    for f in xlsx_files:
        update_xlsx(f)

    print("\n[COMPLETE] All documents updated.")

if __name__ == "__main__":
    main()
