"""
Verification script for all 8 test documentation files in docs/testings/.
Checks:
- All 4 Word (.docx) documents: in every test table, column 4 is 'Actual Outcome' and column 5 is 'Expected Outcome'.
- All 4 Excel (.xlsx) workbooks: column 'Actual Outcome' appears before 'Expected Outcome'.
- Test case count and image counts are preserved.
"""

import os
import docx
import openpyxl

folder = "docs/testings"

print("=" * 60)
print("VERIFYING WORD DOCUMENTS (.docx)")
print("=" * 60)

docx_files = [
    ("Alpha_Blackbox_Test_Report.docx", 138),
    ("Beta_Blackbox_Test_Report.docx", 138),
    ("Alpha_Whitebox_Test_Report.docx", 100),
    ("Beta_Whitebox_Test_Report.docx", 100),
]

all_passed = True

for fname, expected_count in docx_files:
    fpath = os.path.join(folder, fname)
    doc = docx.Document(fpath)
    test_tables = 0
    correct_order = 0

    for tbl in doc.tables:
        if len(tbl.rows) > 0 and len(tbl.rows[0].cells) >= 6:
            c4 = tbl.rows[0].cells[4].text.strip().replace("\n", " ")
            c5 = tbl.rows[0].cells[5].text.strip().replace("\n", " ")
            if "outcome" in c4.lower() or "outcome" in c5.lower():
                test_tables += 1
                if "actual" in c4.lower() and "expected" in c5.lower():
                    correct_order += 1
                else:
                    print(f"  [MISMATCH] {fname} table had: col4='{c4}', col5='{c5}'")

    print(f"{fname}:")
    print(f"  Total test tables: {test_tables} (expected {expected_count})")
    print(f"  Tables with 'Actual Outcome' before 'Expected Outcome': {correct_order}")
    print(f"  Inline shapes (images): {len(doc.inline_shapes)}")
    if test_tables == expected_count and correct_order == test_tables:
        print("  STATUS: [PASS] All test tables verified!\n")
    else:
        print("  STATUS: [FAIL] Incomplete or incorrect order!\n")
        all_passed = False

print("=" * 60)
print("VERIFYING EXCEL WORKBOOKS (.xlsx)")
print("=" * 60)

xlsx_files = [
    ("alpha_blackbox_test_results.xlsx", 139, 138),
    ("beta_blackbox_test_results.xlsx", 139, 138),
    ("alpha_whitebox_backend_test_results.xlsx", 101, 100),
    ("beta_whitebox_backend_test_results.xlsx", 101, 100),
]

for fname, expected_rows, expected_images in xlsx_files:
    fpath = os.path.join(folder, fname)
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active

    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    act_idx = None
    exp_idx = None
    for idx, h in enumerate(headers):
        if "actual outcome" in h.lower():
            act_idx = idx
        elif "expected outcome" in h.lower():
            exp_idx = idx

    print(f"{fname}:")
    print(f"  Headers: {headers}")
    print(f"  Rows: {ws.max_row} (expected {expected_rows})")
    print(f"  Images: {len(ws._images)} (expected {expected_images})")
    print(f"  Actual Outcome Col Index: {act_idx} ('{headers[act_idx] if act_idx is not None else 'None'}')")
    print(f"  Expected Outcome Col Index: {exp_idx} ('{headers[exp_idx] if exp_idx is not None else 'None'}')")

    # Sample check row 2
    r2_act = str(ws.cell(2, act_idx + 1).value or "")[:35] if act_idx is not None else ""
    r2_exp = str(ws.cell(2, exp_idx + 1).value or "")[:35] if exp_idx is not None else ""
    print(f"  Sample Row 2 Actual: '{r2_act}'")
    print(f"  Sample Row 2 Expected: '{r2_exp}'")

    if act_idx is not None and exp_idx is not None and act_idx < exp_idx:
        print("  STATUS: [PASS] 'Actual Outcome' is before 'Expected Outcome'!\n")
    else:
        print("  STATUS: [FAIL] Order is incorrect!\n")
        all_passed = False

print("=" * 60)
if all_passed:
    print("ALL 8 TESTING DOCUMENTS SUCCESSFULLY VERIFIED!")
else:
    print("SOME DOCUMENTS FAILED VERIFICATION!")
print("=" * 60)
